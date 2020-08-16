from QCD_test.http_request import http_request
from QCD_test.R_Wtest_case import R_Wtest_case
all_case=R_Wtest_case('test_case.xlsx','recharge')
from openpyxl import load_workbook
# 提取token
def run():
    global Token
    for i in range(len(all_case)):
        print(all_case[i])
        if i == 0:
            url = all_case[0][6]
            param = eval(all_case[0][7])
            header = eval(all_case[0][3])
            response = http_request(url, param, header)  # 发送一个登录请求
            print(response)
            Token = response['data']['token_info']['token']
        else:
            url = all_case[i][6]
            param = eval(all_case[i][7])
            header = eval(all_case[i][3])
            header['Authorization']='Bearer '+Token
            response=http_request(url,param,header)
            print(response)
        wb=load_workbook('test_case.xlsx')
        sheet=wb['recharge']
        value=sheet.cell(row=i+2,column=11).value=str(response)
        actual={'code':response['code'],'msg':response['msg']}
        excepted=eval(all_case[i][8])
        if actual==excepted:
            print('测试通过')
            sheet.cell(row=i+2,column=12).value='pass'
        else:
            print('测试不通过')
            sheet.cell(row=i + 2, column=12).value = 'fail'
        wb.save('test_case.xlsx')
run()