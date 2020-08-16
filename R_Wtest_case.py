from openpyxl import load_workbook
def R_Wtest_case(workbook,sheet):
    all_case=[]
    wb=load_workbook(workbook)
    sheet_rec=wb[sheet]
    for i in range(1,sheet_rec.max_row):
        case=[]
        for j in range(sheet_rec.max_column):
            value=sheet_rec.cell(row=i+1,column=j+1).value
            case.append(value)
        all_case.append(case)
    return all_case
if __name__ == '__main__':

    all_case=R_Wtest_case('test_case.xlsx','recharge')